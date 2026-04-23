"""
File parsing utilities for document processing.
"""
import os
import re
from pathlib import Path


def extract_text(file_path: str) -> str:
    """
    Extract text from a document based on file type.
    Routes to appropriate parser.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    
    if ext == '.pdf':
        return extract_text_from_pdf(file_path)
    elif ext == '.docx':
        return extract_text_from_docx(file_path)
    elif ext == '.xlsx':
        return extract_text_from_xlsx(file_path)
    elif ext in ['.jpg', '.jpeg', '.png']:
        return extract_text_from_image(file_path)
    elif ext == '.txt':
        return extract_text_from_txt(file_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from PDF using PyMuPDF."""
    try:
        try:
            import fitz  # PyMuPDF
        except ImportError as ie:
            raise ValueError("PyMuPDF (fitz) import failed. Please check installation. Error: {}".format(str(ie)))
        doc = fitz.open(file_path)
        text = ''
        for page in doc:
            text += page.get_text()
        if not text.strip():
            raise ValueError("No text extracted from PDF. PDF may be scanned or empty.")
        return text
    except Exception as e:
        # Log error and fallback to OCR if digital PDF extraction fails
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"PDF extraction failed for {file_path}: {str(e)}")
        return extract_text_from_image_ocr(file_path)


def extract_text_from_docx(file_path: str) -> str:
    """Extract text from DOCX using python-docx."""
    try:
        from docx import Document as DocxDocument
        doc = DocxDocument(file_path)
        text = '\n'.join([para.text for para in doc.paragraphs])
        return text
    except Exception as e:
        raise ValueError(f"Error extracting DOCX: {str(e)}")


def extract_text_from_xlsx(file_path: str) -> str:
    """Extract text from ALL sheets in xlsx."""
    import openpyxl
    wb         = openpyxl.load_workbook(file_path, data_only=True)
    text_parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f'Sheet: {sheet_name}')

        for row in ws.iter_rows(values_only=True):
            if all(cell is None for cell in row):
                continue
            row_text = '\t'.join(
                str(cell).strip() if cell is not None else ''
                for cell in row
            )
            if row_text.strip():
                text_parts.append(row_text)

        text_parts.append('')

    return '\n'.join(text_parts)


def extract_text_from_image(file_path: str) -> str:
    """Extract text from image using Tesseract OCR."""
    try:
        return extract_text_from_image_ocr(file_path)
    except Exception as e:
        raise ValueError(f"Error extracting image: {str(e)}")


def extract_text_from_image_ocr(file_path: str) -> str:
    """Extract text from image using Tesseract OCR."""
    try:
        import pytesseract
        from PIL import Image
        image = Image.open(file_path)
        text = pytesseract.image_to_string(image)
        if text.strip():
            return text
        else:
            raise ValueError("No text could be extracted from this image")
    except ImportError:
        raise ValueError("pytesseract not installed. Please install: pip install pytesseract")
    except Exception as e:
        error_msg = str(e)
        if 'tesseract' in error_msg.lower() or 'not found' in error_msg.lower():
            raise ValueError(f"Tesseract OCR is not installed on this system. Error: {error_msg}\n"
                           "Install Tesseract from: https://github.com/UB-Mannheim/tesseract/wiki")
        raise ValueError(f"Image OCR failed: {error_msg}")


def extract_text_from_txt(file_path: str) -> str:
    """Extract text from plain text file."""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        raise ValueError(f"Error reading text file: {str(e)}")


def scrub_pii(text: str) -> str:
    """
    Remove personally identifiable information from text.
    Scrubs: Aadhaar, PAN, phone numbers, bank accounts.
    """
    # Aadhaar (12 digits)
    text = re.sub(r'\b\d{4}\s?\d{4}\s?\d{4}\b', '[AADHAAR_REMOVED]', text)
    
    # PAN (10 chars: 5 letters, 4 digits, 1 letter)
    text = re.sub(r'\b[A-Z]{5}[0-9]{4}[A-Z]{1}\b', '[PAN_REMOVED]', text)
    
    # Indian phone numbers (10 digits, various formats)
    text = re.sub(r'\b(?:\+91|0)?[789]\d{9}\b', '[PHONE_REMOVED]', text)
    
    # Bank account numbers (8-17 digits)
    text = re.sub(r'\b\d{8,17}\b', '[ACCOUNT_REMOVED]', text)
    
    # Credit/Debit card numbers
    text = re.sub(r'\b\d{4}[\s-]?\d{4}[\s-]?\d{4}[\s-]?\d{4}\b', '[CARD_REMOVED]', text)
    
    return text


def chunk_text(text: str, chunk_size: int = 500, overlap: int = 50) -> list:
    """
    Split text into overlapping chunks for embedding.
    chunk_size: approximate tokens per chunk (500 ~ 2000 chars)
    overlap: approximate tokens overlap between chunks (50 ~ 200 chars)
    """
    words = text.split()
    chunks = []
    current_chunk = []
    
    for word in words:
        current_chunk.append(word)
        # Approximate: 1 token ≈ 4 chars, so 500 tokens ≈ 2000 chars
        if len(' '.join(current_chunk)) >= chunk_size * 4:
            chunk_text = ' '.join(current_chunk)
            chunks.append(chunk_text)
            
            # Create overlap
            overlap_words = int(overlap / 4) // len(current_chunk[-1:]) if current_chunk else 0
            current_chunk = current_chunk[-max(5, overlap_words):] if len(current_chunk) > 5 else current_chunk
    
    # Add remaining text as final chunk
    if current_chunk:
        chunks.append(' '.join(current_chunk))
    
    return chunks
